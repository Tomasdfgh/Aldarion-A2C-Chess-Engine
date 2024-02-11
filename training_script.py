import chess
import chess.svg
import cairosvg
import matplotlib.pyplot as plt
from io import BytesIO
import model as md
import board_reader as br
import torch
import board_display as bd
import MTCS as mt
import torch.optim as optim
import os
import openpyxl

model = md.ChessNet().double()
model.train()

gamma = 0.99

beta = 0.15

optimizer = optim.AdamW(model.parameters(), lr = 0.001)

torch.autograd.set_detect_anomaly(True)

#fig2, ax2 = bd.get_fig_ax()

#Load up the workbook
workbook = openpyxl.load_workbook('DataSet.xlsx')
sheet = workbook.active
if sheet.max_row > 1:
    game_num = sheet[sheet.max_row][0].value
else:
    game_num = 0

# board = chess.Board('8/8/2bb4/5pkp/2p5/2P4K/8/8 w - - 2 68')
# board.turn = not board.turn
# board.push_uci('f5f4')
# board.turn = not board.turn
# board.push_uci('f4f3')
# board.turn = not board.turn
# board.push_uci('f3f2')
# board.turn = not board.turn
# print(board.fen())
# fig, ax = bd.get_fig_ax()
# while True:
#     bd.render_board(board, fig, ax)


if __name__ == "__main__":
    path_taken = mt.run_game(model)

    for i in path_taken:
        sheet.append(i)
    workbook.save('DataSet.xlsx')
    pass
